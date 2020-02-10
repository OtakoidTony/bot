import discord
import shutil
import sqlite3
import smtplib
from email.mime.text import MIMEText
import sys
import wikipediaapi
from google_images_download import google_images_download
import ssl
import asyncio
from googletrans import Translator
import random
from discord import Webhook, RequestsWebhookAdapter
from urllib.request import urlopen, Request
from discord.utils import get
import bs4
import requests
import openpyxl
import youtube_dl
import urllib.request
import os
from bs4 import BeautifulSoup
import urllib
import datetime
con=sqlite3.connect('test.db')
cur=con.cursor()
queues={}
token=os.environ["TOKEN"]
client=discord.Client()
admin='647630912795836437'
black=[0]
vip=[0]
print(f'키키봇을 시작하는중입니다...몇초가 걸릴수 있습니다\n\n\n')
afsdaf=cur.execute('SELECT * FROM USERS')
for i in afsdaf:
    if i[4]==1:
        vip.append(i[0])
    if i[5]==1:
        black.append(i[0])
@client.event
async def on_ready():
    print('----------------------------------------------------------------------------')
    print(f'| Client ID: {client.user.id}                                             |')
    print(f'| Client Name: {client.user.name}                                                     |')
    print(f'| Client Token: {str(token)} |')
    print('----------------------------------------------------------------------------\n\n=========================================\n\n')  
@client.event
async def on_message(message):
    async def makeembed(title, description):
        now=datetime.datetime.now()
        embed=discord.Embed(
            title=title,
            description=description,
            colour=discord.Colour.blue()
        )
        embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
        await message.channel.send(embed=embed)
    try:
        if message.author.bot:
            return None
        if message.content.startswith('키키야'):
            print(f'{message.author.display_name}님의 메시지: {message.content}')
            if str(message.author.id) in black:
                await makeembed('당신은 키키봇 블랙리스트 목록에 있습니다', '블랙리스트 감지 기능')
                return 
            if message.content=='키키야': 
                now=datetime.datetime.now()
                msg='안녕하세여^^'
                embed=discord.Embed(
                    title=str(msg),
                    description="챗 기능",
                    colour=discord.Colour.blue()
                )
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 실시간검색어') or message.content.startswith('키키야 실검'):
                json=requests.get("https://www.naver.com/srchrank?frm=main").json()
                embed = discord.Embed(
                    title='네이버 실시간 검색어',
                    description='실시간검색어',
                    colour=discord.Colour.blue()
                )
                for r in json.get('data'):
                    a=r.get("keyword")
                    embed.add_field(name=f'{r.get("rank")}위', value=r.get("keyword"), inline=False)
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            elif message.content=='키키야 섭정보' or message.content=='키키야 서버정보':
                now=datetime.datetime.now()
                embed=discord.Embed(colour=discord.Color.blue())
                embed.set_author(name=f'서버 정보 - {message.guild.name}')
                embed.set_thumbnail(url=message.guild.icon_url)
                embed.add_field(name="이름", value=message.guild.name, inline=False)
                embed.add_field(name="아이디", value=message.guild.id, inline=False)
                embed.add_field(name="나라", value=str(message.guild.region).title(), inline=False)
                embed.add_field(name="주인", value=message.guild.owner.display_name, inline=False)
                embed.add_field(name="멤버 수", value=len(message.guild.members), inline=False)
                embed.add_field(name="생긴 날짜", value=message.guild.created_at.strftime("20%y년 %m월 %d일"), inline=False)
                #embed.add_field(name="최근에 들어온 유저", value=[Member for Member in Guild.members if Member.joined_at is max([Member.joined_at for Member in get.guild.member])], inline=False)
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 디비조회'):
                if str(message.author.id)!=str(admin):
                    await makeembed('해킹 ㄴㄴ','관리자 기능')
                    return
                db=message.content.split(' ')[2]
                if db=='서버':
                    db='SERVERS'
                elif db=='유저':
                    db='USERS'
                elif db=='기억':
                    db='CUSTOMCOMMANDS'
                elif db=='도움말':
                    db='HELPS'
                else:
                    await makeembed('그런 데이터베이스는 없습니다', '에러 감지 기능')
                    return
                dsaf=""
                cur.execute(f'SELECT * FROM {db}')
                for row in cur:
                    dsaf+=f'{str(row)}\n'
                await makeembed(f'테이블 {db} 값', dsaf)
            elif message.content.startswith("키키야 날씨"):
                learn = message.content.split(" ")
                location = learn[2]
                enc_location = urllib.parse.quote(location+'날씨')
                hdr = {'User-Agent': 'Mozilla/5.0'} 
                url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + enc_location
                req = Request(url, headers=hdr)
                html = urllib.request.urlopen(req)
                bsObj = bs4.BeautifulSoup(html, "html.parser")
                todayBase = bsObj.find('div', {'class': 'main_info'})
                todayTemp1 = todayBase.find('span', {'class': 'todaytemp'})
                todayTemp = todayTemp1.text.strip()  # 온도
                todayValueBase = todayBase.find('ul', {'class': 'info_list'})
                todayValue2 = todayValueBase.find('p', {'class': 'cast_txt'})
                todayValue = todayValue2.text.strip()  
                todayFeelingTemp1 = todayValueBase.find('span', {'class': 'sensible'})
                todayFeelingTemp = todayFeelingTemp1.text.strip()  
                todayMiseaMongi1 = bsObj.find('div', {'class': 'sub_info'})
                todayMiseaMongi2 = todayMiseaMongi1.find('div', {'class': 'detail_box'})
                todayMiseaMongi3 = todayMiseaMongi2.find('dd')
                todayMiseaMongi = todayMiseaMongi3.text  
                tomorrowBase = bsObj.find('div', {'class': 'table_info weekly _weeklyWeather'})
                tomorrowTemp1 = tomorrowBase.find('li', {'class': 'date_info'})
                tomorrowTemp2 = tomorrowTemp1.find('dl')
                tomorrowTemp3 = tomorrowTemp2.find('dd')
                tomorrowTemp = tomorrowTemp3.text.strip()  
                tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
                tomorrowMoring1 = tomorrowAreaBase.find('div', {'class': 'main_info morning_box'})
                tomorrowMoring2 = tomorrowMoring1.find('span', {'class': 'todaytemp'})
                tomorrowMoring = tomorrowMoring2.text.strip()  
                tomorrowValue1 = tomorrowMoring1.find('div', {'class': 'info_data'})
                tomorrowValue = tomorrowValue1.text.strip()  
                tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
                tomorrowAllFind = tomorrowAreaBase.find_all('div', {'class': 'main_info morning_box'})
                tomorrowAfter1 = tomorrowAllFind[1]
                tomorrowAfter2 = tomorrowAfter1.find('p', {'class': 'info_temperature'})
                tomorrowAfter3 = tomorrowAfter2.find('span', {'class': 'todaytemp'})
                tomorrowAfterTemp = tomorrowAfter3.text.strip() 
                tomorrowAfterValue1 = tomorrowAfter1.find('div', {'class': 'info_data'})
                tomorrowAfterValue = tomorrowAfterValue1.text.strip()
                embed = discord.Embed(
                    title=learn[1]+ ' 정보',
                    description=learn[1]+ ' 정보입니다.',
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                embed.add_field(name='현재온도', value=todayTemp+'˚', inline=False)  # 현재온도
                embed.add_field(name='체감온도', value=todayFeelingTemp, inline=False)  # 체감온도
                embed.add_field(name='현재상태', value=todayValue, inline=False)  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
                embed.add_field(name='현재 미세먼지 상태', value=todayMiseaMongi, inline=False)  # 오늘 미세먼지
                embed.add_field(name='오늘 오전/오후 날씨', value=tomorrowTemp, inline=False)  # 오늘날씨 # color=discord.Color.blue()
                embed.add_field(name='**----------------------------------**',value='**----------------------------------**', inline=False)  # 구분선
                embed.add_field(name='내일 오전온도', value=tomorrowMoring+'˚', inline=False)  # 내일오전날씨
                embed.add_field(name='내일 오전날씨상태, 미세먼지 상태', value=tomorrowValue, inline=False)  # 내일오전 날씨상태
                embed.add_field(name='내일 오후온도', value=tomorrowAfterTemp + '˚', inline=False)  # 내일오후날씨
                embed.add_field(name='내일 오후날씨상태, 미세먼지 상태', value=tomorrowAfterValue, inline=False)  # 내일오후 날씨상태
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 단축'):
                target=message.content.split(' ')[2]
                client_id="DqTSCjayP8uFjYJCWA3r"
                client_secret="KsaviRkocB"
                header = {'X-Naver-Client-Id': client_id, 'X-Naver-Client-Secret': client_secret}
                naver = 'https://openapi.naver.com/v1/util/shorturl'
                data = {'url': target}
                maker=requests.post(url=naver,data=data,headers=header)
                maker.close()
                output=maker.json()['result']['url']
                await makeembed(f'단축 주소: {output}', '응용 기능')
            elif message.content.startswith('키키야 블랙추가'):
                if str(message.author.id)!=str(admin):
                    await makeembed('블랙리스트 안에 들고 싶어요?!', '관리자 기능')
                idd=message.content.split(' ')[2]
                asdfd=cur.execute('SELECT * FROM USERS')
                for row in asdfd:
                    if str(message.author.id) in row:
                        cur.execute(f'UPDATE USERS SET black=1 WHERE id={idd}')
                        await makeembed('유저를 블랙리스트 등록에 성공하였습니다', '관리자 기능')
                        black.append(idd)
                        con.commit()
                        return
                await makeembed('다행히 이 사람은 키키봇 서비스에 가입이 되어 있지 않네요', '관리자 기능')
            elif message.content.startswith('키키야 블랙해제'):
                if str(message.author.id)!=str(admin):
                    await makeembed('블랙리스트 안에 들고 싶어요?!', '관리자 기능')
                idd=message.content.split(' ')[2]
                asdfd=cur.execute('SELECT * FROM USERS')
                for row in asdfd:
                    if str(message.author.id) in row:
                        cur.execute(f'UPDATE USERS SET black=0 WHERE id={idd}')
                        await makeembed('유저를 블랙리스트 해제에 성공하였습니다', '관리자 기능')
                        black.append(idd)
                        con.commit()
                        return
                await makeembed('다행히(?) 이 사람은 키키봇 서비스에 가입이 되어 있지 않네요', '관리자 기능')
            elif message.content.startswith('키키야 공지') and message.content.startswith('키키야 공지등록')==False:
                asdfdsa=cur.execute('SELECT * FROM SERVERS')
                if message.author.id!=647630912795836437:
                    await makeembed('키키 해킹하고 다시 오세여', '관리자 기능')
                    return None
                msg=message.content[7:]
                now=datetime.datetime.now()
                embed=discord.Embed(
                    title=msg.split('$')[0],
                    description=msg.split('$')[1],
                    colour=discord.Colour.blue()
                ).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                for i in client.guilds:
                    arr=[0]
                    alla=False
                    flag=True
                    z=0
                    for row in asdfdsa:
                        if str(row[0])==str(i.id):
                            try:
                                await client.get_channel(int(row[2])).send(embed=embed)
                            except:
                                flag=False
                                break
                    if flag==False:
                        continue
                    for j in i.channels:
                        arr.append(j.id)
                        z+=1
                        if '📢봇_공지' in j.name or '봇-공지' in j.name or '키키봇' in j.name or '아파트-봇-공지사항' in j.name:
                            if str(j.type)=='text':
                                try:
                                    await j.send(embed=embed)
                                    alla=True
                                except:
                                    pass
                                break
                    if alla==False:
                        try:
                            chan=i.channels[1]
                        except:
                            pass
                        if str(chan.type)=='text':
                            try:
                                await chan.send(embed=embed)
                            except:
                                pass
                await makeembed('공지 전송을 완료했습니다', '관리자 기능')
            elif message.content=='키키야 영화순위':
                url = urlopen("https://movie.naver.com/movie/running/current.nhn")
                bs = BeautifulSoup(url, 'html.parser')
                body = bs.body
                now=datetime.datetime.now()
                target = body.find(class_="lst_detail_t1")
                embed=discord.Embed(title="영화 순위", description="네이버 영화에서 크롤링", colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                list = target.find_all('li')
                no = 1
                for n in range(0, 9) :
                    no += 1
                    title = list[n].find(class_="tit").find("a").text
                    try:
                        director = list[n].find(class_="info_txt1").find_all("dd")[1].find("span").find_all("a")
                        directorList = [director.text.strip() for director in director]
                    except IndexError:
                        directorList="정보 없음"
                    try:
                        cast = list[n].find(class_="lst_dsc").find("dl", class_="info_txt1").find_all("dd")[2].find(class_="link_txt").find_all("a")
                        castList = [cast.text.strip() for cast in cast]
                    except IndexError:
                        castList="정보 없음"
                    embed.add_field(name=f'{no}등', value=f"영화 제목:  {title}\n제작 감독:  {directorList}\n출연 배우:  {castList}", inline=False)
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 번역'):
                translator = Translator()
                index=message.content[7:]
                asdf='en'
                await makeembed(f'{index}를 {translator.detect(index).lang}에서 영어로 변경: \n{translator.translate(index, src=translator.detect(index).lang, dest=asdf).text}', '응용 기능')
            elif message.content.startswith('키키야 내정보'):
                now=datetime.datetime.now()
                roles=[role for role in message.author.roles]
                embed=discord.Embed(colour=discord.Color.blue())
                embed.set_author(name=f"유저정보 - {message.author.display_name}")
                embed.set_thumbnail(url=message.author.avatar_url)
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                embed.add_field(name="아이디", value=message.author.id, inline=False)
                #embed.add_field(name="이메일", value=message.author.email, inline=False)
                embed.add_field(name="계정 생성 시간", value=message.author.created_at.strftime("%a, %#d %B %Y, %I:%M %p UTC"), inline=False)
                embed.add_field(name="가입 시간", value=message.author.joined_at.strftime("%a, %#d %B %Y, %I:%M %p UTC"), inline=False)
                embed.add_field(name=f"가진 역할들 ({len(roles)})", value=" ".join([role.mention for role in roles]), inline=False)
                embed.add_field(name="가장 높은 역할", value=message.author.top_role.mention, inline=False)
                embed.add_field(name="상태", value=message.author.status, inline=False)
                embed.add_field(name="봇", value=message.author.bot, inline=False)
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 차단'):
                if message.author.id==647630912795836437:
                    author=message.guild.get_member(int(message.content[7:].split('/')[0]))
                    why=message.content[7:].split('/')[1]
                    await message.guild.ban(author)
                    now=datetime.datetime.now()
                    await author.send(embed=discord.Embed(title=f'{message.guild} 에서 차단당했습니다', description=f'이유: {str(why)}', colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일'))
                    await makeembed('성공적으로 차단 완료!', '관리자 기능')
                else:
                    await makeembed('권한이나 가지고 다시 오세여', '에러 감지 기능')
            elif message.content.startswith('키키야 도움말추가'):
                if message.author.id!=647630912795836437:
                    await makeembed('키키봇 해킹하고 다시 오세여', '관리자 기능')
                    return None
                file3=openpyxl.load_workbook('도움말.xlsx')
                sheet3=file3.active
                strr=message.content[10:]
                lea=strr.split('/')
                i=1
                while sheet3['A'+str(i)].value!=None:
                    i+=1
                sheet3['A'+str(i)].value=str(lea[0]) #형식: 관리자, 돈, 채팅, 응용
                sheet3['B'+str(i)].value=str(lea[1]) #명령어
                sheet3['C'+str(i)].value=str(lea[2]) #설명
                file3.save('도움말.xlsx')
                await makeembed('키키봇 도움말 추가 성공!', '관리자 기능')
            elif message.content.startswith('키키야 엑셀초기화'):
                if message.author==admin:
                    await makeembed('응 안되', '관리자 기능')
                    return
                file=openpyxl.load_workbook(message.content.split(' ')[2])
                sheet=file.active
                j=1
                while sheet['A'+str(j)].value!=None:
                    i=65
                    while sheet[chr(i)+str(j)].value!=None:
                        sheet[chr(i)+str(j)].value=None
                        i+=1
                    j+=1
                file.save(message.content.split(' ')[2])
                await makeembed('엑셀파일 초기화 완료', '관리자 기능')
            elif message.content.startswith('키키야 공지등록'):
                asdfdsa=cur.exectue('SELECT * FROM SERVERS')
                if message.content[9:]!=None:
                    chann=str(message.content[9:])
                else:
                    chann=str(message.channel.id)
                for row in asdfdsa:
                    if str(message.author.id) in row:
                        cur.execute(f"INSERT INTO SERVERS Values(?, ?, ?, ?);", (str(message.guild.id), message.guild.name, str(chann), str(message.guild.owner.id)))
                        con.commit()
                        await makeembed('키키봇 공지채널 설정 완료!', '응용 기능')
                        return None
                cur.execute(f"INSERT INTO SERVERS Values(?, ?, ?, ?);", (str(message.guild.id), message.guild.name, str(chann), str(message.guild.owner.id)))
                await makeembed('키키봇 공지채널 설정 완료!', '응용 기능')
                con.commit()
            elif message.content.startswith('키키야 실험'):
                if message.author.id!=647630912795836437:
                    await makeembed('관리자 해킹하고 오라니까여', '관리자 기능')
                    return None
                exect=message.content[7:]
                eval(exect)
            elif message.content=='키키야 도움말' or message.content=='키키야 도움':
                embed=discord.Embed(
                    title="키키봇 도움말",
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                embed.add_field(name="키키야 관리자도움말", value="관리자 전용 도움말입니다(볼순 있지만 쓸순 없어요^^)", inline=False)
                embed.add_field(name="키키야 응용도움말", value="대부분 크롤링을 이용한 기능이죠^^", inline=False)
                embed.add_field(name="키키야 돈도움말", value="돈과 상점을 사용할수 있는 명령어들이죠^^", inline=False)
                embed.add_field(name="키키야 채팅도움말", value="커스텀명령어와 커스텀링크를 활용하는 기능들이죠^^", inline=False)
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 말해'):
                talk=message.content[7:]
                await message.channel.purge(limit=1)
                if message.content=='키키야 말해 키키천재':
                    await makeembed('축하합니다^^\n이스터에그를 찾으셨어요\n||(근데 할수 있는게 없어서 쓸모없는 이스터애그)||', '챗 기능')
                    return None
                try:
                    if talk.split('|')[1]=='X':
                        await message.channel.send(talk.split('|')[0])
                    elif talk.split('|')[1]!='X':
                        await makeembed('도데체 | 를 썼으면 뒤에다 X를 써야죠 뭐를 쓰려 한거에여?!', '에러 감지 기능')
                except IndexError:
                    await makeembed(talk, '챗 기능')
            elif message.content=='키키야 멜론차트':
                if __name__=="__main__":
                    RANK=10
                    header={'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Trident/7.0; rv:11.0) like Gecko'}
                    req = requests.get('https://www.melon.com/chart/index.htm', headers = header)
                    html = req.text
                    parse = BeautifulSoup(html, 'html.parser')
                    titles = parse.find_all("div", {"class": "ellipsis rank01"})
                    songs = parse.find_all("div", {"class": "ellipsis rank02"})
                    title = []
                    song = []
                    embed=discord.Embed(
                        title="맬론차트 1~10위", 
                        color=discord.Colour.green()
                    )
                    for t in titles:
                        title.append(t.find('a').text)
                    for s in songs:
                        song.append(s.find('span', {"class": "checkEllipsis"}).text)
                    for i in range(RANK):
                        embed.add_field(name='%3d위'%(i+1), value='%s - %s'%(title[i], song[i]), inline=False)
                    now=datetime.datetime.now()
                    embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                    await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 건의'):
                learn=message.content[7:]
                author=message.guild.get_member(647630912795836437)
                embed=discord.Embed(
                    title=f"{message.author.display_name}님의 건의사항: {learn}",
                    description="응용 기능",
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await author.send(embed=embed)
                await makeembed('건의사항이 성공적으로 전달되었습니다', '응용 기능')
            elif message.content.startswith('키키야 이메일'):
                email=message.content.split(' ')[2]
                m2sg=message.content.split(email)[1][1:]
                s = smtplib.SMTP('smtp.gmail.com', 587)
                s.starttls()
                print(email)
                s.login('happykiki7000@gmail.com', os.environ["EMAIL"])
                msg = MIMEText(m2sg)
                msg['Subject'] = f'{message.author.display_name}    님의 메시지'
                s.sendmail("happykiki7000@gmail.com", email, msg.as_string())
                s.quit()
                await makeembed('이메일 전송에 성공하였습니다', '응용 기능')
            elif message.content.startswith('키키야 무단침입'):
                strrrrrr=message.content[9:]
                channel=client.get_channel(int(strrrrrr.split(' ')[0]))
                await message.channel.purge(limit=1)
                now=datetime.datetime.now()
                await channel.send(embed=discord.Embed(title=str(strrrrrr[19:]), description='응용 기능', colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일'))
            elif message.content.startswith('키키야 디엠전송'):
                strrr=message.content[9:]
                msg=strrr[19:]
                author=message.guild.get_member(int(strrr.split(' ')[0]))
                embed=discord.Embed(
                    title=f"{msg}",
                    description="응용 기능",
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await author.send(embed=embed)
                await makeembed('디엠으로 메시지가 성공적으로 전달되었습니다', '응용 기능')
            elif message.content=='키키야 가입':
                cur.execute("SELECT * FROM USERS")
                for row in cur:
                    if str(message.author.id) in row:
                        await makeembed('이미 가입된 사용자입니다', '에러 감지 기능')
                        return
                cur.execute(f"INSERT INTO USERS Values(?, ?, ?, ?, ?, ?, ?);", (str(message.author.id), message.author.display_name, str(0), str(0), str(0), str(0), str(0)))
                con.commit()
                await makeembed('키키봇 서비스 가입 완료!', '챗 기능')
            elif message.content=='키키야 핑':
                embed = discord.Embed(
                    title = 'PONG!',
                    author = '키키봇',
                    colour = discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                embed.add_field(name = 'Result', value = f' {round(client.latency * 1000)}ms' + f' (0.{round(client.latency * 1000)}초)', inline = False)
                await message.channel.send(embed = embed)
            elif message.content.startswith('키키야 찬반투표'):
                lern=message.content[9:]
                embed=discord.Embed(
                    title=str(lern),
                    description=f"{message.author.display_name}님의 찬반투표",
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                msg=await message.channel.send(embed=embed)
                await msg.add_reaction('👍')
                await msg.add_reaction('👎')
            elif message.content=='키키야 봇정보':
                saf=""
                for i in client.guilds:
                    saf+=str(i)
                    saf+='\n'
                embed=discord.Embed(
                    title="키키봇 정보",
                    description=f"현재 {len(client.guilds)}개의 서버에 참여중, {len(client.users)}명의 유저들과 소통중\n\n**키키봇이 참여하고 있는 서버**\n\n\n{saf}",
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            elif message.content=='키키야 관리자도움말':
                file3=openpyxl.load_workbook('도움말.xlsx')
                sheet3=file3.active
                i=1
                embed=discord.Embed(
                    title="키키봇 관리자도움말",
                    description='키키봇 관리자용 도움말입니다',
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                embed.add_field(name='키키야 도움말추가 [형식]/[커맨드]/[설명]', value='키키봇의 도움말을 추가하는 기능입니다', inline=False)
                embed.add_field(name='키키야 공지 [제목]/[내용]', value='키키봇 공지 기능입니다', inline=False)
                while sheet3['A'+str(i)].value!=None:
                    if sheet3['A'+str(i)].value=='관리자':
                        embed.add_field(name=str(sheet3['B'+str(i)].value), value=str(sheet3['C'+str(i)].value), inline=False)
                    i+=1
                await message.channel.send(embed=embed)
            elif message.content=='키키야 응용도움말':
                file3=openpyxl.load_workbook('도움말.xlsx')
                sheet3=file3.active
                i=1
                embed=discord.Embed(
                    title="키키봇 응용도움말",
                    description='키키봇 응용 도움말입니다',
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                while sheet3['A'+str(i)].value!=None:
                    if sheet3['A'+str(i)].value=='응용':
                        embed.add_field(name=str(sheet3['B'+str(i)].value), value=str(sheet3['C'+str(i)].value), inline=False)
                    i+=1
                await message.channel.send(embed=embed)
            elif message.content=='키키야 돈도움말':
                file3=openpyxl.load_workbook('도움말.xlsx')
                sheet3=file3.active
                i=1
                embed=discord.Embed(
                    title="키키봇 돈도움말",
                    description='키키봇 돈 관련 도움말입니다',
                    colour=discord.Colour.blue()
                )
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                while sheet3['A'+str(i)].value!=None:
                    if sheet3['A'+str(i)].value=='돈':
                        embed.add_field(name=str(sheet3['B'+str(i)].value), value=str(sheet3['C'+str(i)].value), inline=False)
                    i+=1
                await message.channel.send(embed=embed)
            elif message.content=='키키야 채팅도움말':
                file3=openpyxl.load_workbook('도움말.xlsx')
                sheet3=file3.active 
                i=1
                embed=discord.Embed(
                    title="키키봇 채팅도움말",
                    description='키키봇 채팅용 도움말입니다',
                    colour=discord.Colour.blue()
                )
                embed.add_field(name='키키야 커명추가 [입력]/[출력]', value='커스텀명령어를 추가합니다', inline=False)
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                while sheet3['A'+str(i)].value!=None:
                    if sheet3['A'+str(i)].value=='채팅':
                        embed.add_field(name=str(sheet3['B'+str(i)].value), value=str(sheet3['C'+str(i)].value), inline=False)
                    i+=1
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 청소'):
                varrr=message.content.split(' ')
                await message.channel.purge(limit=int(varrr[2])+1)
                now=datetime.datetime.now()
                msg=await message.channel.send(embed=discord.Embed(title=f'메시지 {str(int(varrr[2]))}개 삭제 완료!', descirption='응용 기능', colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일'))
                await asyncio.sleep(2.5)
                await msg.delete()
            elif message.content.startswith('키키야 한단어커명추가'):
                strsss=message.content[12:]
                i = 1
                if strsss=='도움말' or strsss=='도움':
                    now=datetime.datetime.now()
                    msg='커스텀명령어'
                    embed=discord.Embed(
                       title='키키봇 커스텀명령어',
                       description="키키봇 한단어커스텀명령어 사용하는 방법입니다^^",
                       colour=discord.Colour.blue()
                    )
                    embed.add_field(name="형식", value="키키야 커명추가 [입력] [출력]")
                    embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                    await message.channel.send(embed=embed)
                    return None
                q=strsss.split(" ")[0] #A부분
                a=strsss.split(" ")[1] #B부분
                asdfdsa=cur.execute('SELECT * FROM USERS')
                for row in asdfdsa:
                    if str(message.author.id) in row:
                        if row[3]<=0:
                            await makeembed('커스텀명령어 티켓이 부족합니다', '`키키야 구입 커스텀명령어`를 이용해 커스텀명령어티켓 10개를 사고 다시 해보세여^^')
                            return None
                        elif row[3]>0:
                            cur.execute(f'UPDATE USERS SET customcommands={row[3]-1}')
                            cur.execute(f"INSERT INTO CUSTOMCOMMANDS Values(?, ?, ?, ?);", (str(q), str(a), str(message.author.id), str(message.author.display_name)))
                            con.commit()
                            await makeembed('커스텀명령어 등록 성공', '채팅 기능')
                            return None
                await makeembed("`키키야 가입`을 입력해 먼저 가입을 하세요", '에러 감지 기능')
            elif message.content.startswith('키키야 커명추가'):
                strsss=message.content[9:]
                if strsss=='도움말' or strsss=='도움':
                    now=datetime.datetime.now()
                    msg='커스텀명령어'
                    embed=discord.Embed(
                       title='키키봇 커스텀명령어',
                       description="키키봇 커스텀명령어 사용하는 방법입니다^^",
                       colour=discord.Colour.blue()
                    )
                    embed.add_field(name="형식", value="키키야 커명추가 [입력]/[출력]")
                    embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                    await message.channel.send(embed=embed)
                    return None
                q=strsss.split("/")[0] #A부분
                a=strsss.split("/")[1] #B부분
                asdfdsa=cur.execute('SELECT * FROM USERS')
                for row in asdfdsa:
                    if str(message.author.id) in row:
                        if row[3]<=0:
                            await makeembed('커스텀명령어 티켓이 부족합니다', '`키키야 구입 커스텀명령어`를 이용해 커스텀명령어티켓 10개를 사고 다시 해보세여^^')
                            return None
                        elif row[3]>0:
                            cur.execute(f'UPDATE USERS SET customcommands={row[3]-1}')
                            cur.execute(f"INSERT INTO CUSTOMCOMMANDS Values(?, ?, ?, ?);", (str(q), str(a), str(message.author.id), str(message.author.display_name)))
                            con.commit()
                            await makeembed('커스텀명령어 등록 성공', '채팅 기능')
                            return None
                await makeembed("`키키야 가입`을 입력해 먼저 가입을 하세요", '에러 감지 기능')
            elif message.content.startswith('키키야 구입'):
                value=message.content.split(' ')[2]
                asdfa=cur.execute('SELECT * FROM USERS')
                for row in asdfa:
                    if str(message.author.id) in row:
                        money=row[2]
                        if value=='커명' or value=='커스텀명령어':
                            if money<5000:
                                await makeembed('돈 더 벌고 오세요', '돈 기능')
                                return 
                            cur.execute(f'UPDATE USERS SET money={money-5000} WHERE id={str(message.author.id)}')
                            cur.execute(f'UPDATE USERS SET customcommands={row[3]+50} WHERE id={str(message.author.id)}')
                            await makeembed('커스텀명령어 티켓 구입 완료', '돈 기능')
                            con.commit()
                            return None
                        elif value=='VIP권' or value=='VIP' or value=='vip' or value=='vip권' or value=='브이아이피':
                            if money<100000:
                                await makeembed('돈 더 벌고 오세요', '돈 기능')
                                return None
                            cur.execute(f'UPDATE USERS SET money={money-100000} WHERE id={str(message.author.id)}')
                            cur.execute(f'UPDATE USERS SET vip=1 WHERE id={str(message.author.id)}')
                            vip.append(str(message.author.id))
                            await makeembed('VIP권 구입 완료', '돈 기능')
                            con.commit()
                            return None
                        else:
                            await makeembed('그런 상품 살 돈으로 니트로나 사세요~', '돈 기능')
                            return
                await makeembed("`키키야 가입`을 입력해 먼저 가입을 하세요", '에러 감지 기능')
            elif message.content=='키키야 인벤':
                asdfa=cur.execute('SELECT * FROM USERS')
                for row in asdfa:
                    if str(message.author.id) in row:
                        await makeembed(f'{message.author.display_name}님의 인벤토리', f'커스텀명령어 티켓: {row[3]}개\nVIP권: {row[4]}개')
                        return None
                await makeembed('`키키야 가입` 하고 오세요', '에러 감지 기능')
            elif message.content=='키키야 내프사' or message.content=='키키야 내아바타':
                embed=discord.Embed(
                    title=f"{message.author.display_name}님의 프사",
                    colour=discord.Colour.blue()
                )
                embed.set_image(url=message.author.avatar_url)
                now=datetime.datetime.now()
                embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 이모지'):
                emoji = [" ꒰⑅ᵕ༚ᵕ꒱ ", " ꒰◍ˊ◡ˋ꒱ ", " ⁽⁽◝꒰ ˙ ꒳ ˙ ꒱◜⁾⁾ ", " ༼ つ ◕_◕ ༽つ ", " ⋌༼ •̀ ⌂ •́ ༽⋋ ",
                    " ( ･ิᴥ･ิ) ", " •ө• ", " ค^•ﻌ•^ค ", " つ╹㉦╹)つ ", " ◕ܫ◕ ", " ᶘ ͡°ᴥ͡°ᶅ ", " ( ؕؔʘ̥̥̥̥ ه ؔؕʘ̥̥̥̥ ) ",
                    " ( •́ ̯•̀ ) ",
                    " •̀.̫•́✧ ", " '͡•_'͡• ", " (΄◞ิ౪◟ิ‵) ", " ˵¯͒ བ¯͒˵ ", " ͡° ͜ʖ ͡° ", " ͡~ ͜ʖ ͡° ", " (づ｡◕‿‿◕｡)づ ",
                    " ´_ゝ` ", " ٩(͡◕_͡◕ ", " ⁄(⁄ ⁄•⁄ω⁄•⁄ ⁄)⁄ ", " ٩(͡ï_͡ï☂ ", " ௐ ", " (´･ʖ̫･`) ", " ε⌯(ง ˙ω˙)ว ",
                    " (っ˘ڡ˘ς) ", "●▅▇█▇▆▅▄▇", "╋╋◀", "︻╦̵̵̿╤──", "ー═┻┳︻▄", "︻╦̵̵͇̿̿̿̿══╤─",
                    " ጿ ኈ ቼ ዽ ጿ ኈ ቼ ዽ ጿ ", "∑◙█▇▆▅▄▃▂", " ♋♉♋ ", " (๑╹ω╹๑) ", " (╯°□°）╯︵ ┻━┻ ",
                    " (///▽///) ", " σ(oдolll) ", " 【o´ﾟ□ﾟ`o】 ", " ＼(^o^)／ ", " (◕‿‿◕｡) ", " ･ᴥ･ ", " ꈍ﹃"                                                                              " ˃̣̣̣̣̣̣︿˂̣̣̣̣̣̣ ",
                    " ( ◍•㉦•◍ ) ", " (｡ì_í｡) ", " (╭•̀ﮧ •́╮) ", " ଘ(੭*ˊᵕˋ)੭ ", " ´_ゝ` ", " (~˘▾˘)~ "] # 이모티콘 배열입니다.
                randomNum = random.randrange(0, len(emoji))
                now=datetime.datetime.now()
                imgembed=discord.Embed(description=emoji[randomNum], colour=discord.Colour.blue())
                imgembed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send( embed=imgembed)
            elif message.content=='키키야 커명확인' or message.content=='키키야 커명목록':
                file = openpyxl.load_workbook("기억.xlsx")
                sheet = file.active
                i=1
                strd=""
                now=datetime.datetime.now()
                while sheet["A" + str(i)].value != None:
                    strd+=f'{str(sheet["A" + str(i)].value)} : {str(sheet["B" + str(i)].value)} \n'
                    i+=1
                embed=discord.Embed(
                    title='명령어들',
                    description=strd,
                    color=discord.Colour.blue()
                ).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.channel.send(embed=embed)
            elif message.content.startswith('키키야 상점'):
                now=datetime.datetime.now()
                embed=discord.Embed(
                    title="키키봇 상점", 
                    description="키키봇 돈으로 살수 있습니다",
                    color=discord.Colour.blue()
                ).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                embed.add_field(name="커스텀명령어 티켓 50장", value="5000원", inline=False)
                embed.add_field(name="VIP권", value="100000원", inline=False)
                await message.channel.send(embed=embed)
            elif message.content.startswith("키키야 입장"):
                global voice
                channel=message.author.voice.channel
                voice=get(client.voice_clients, guild=message.guild)
                if voice and voice.is_connected():
                    await voice.move_to(channel)
                else:
                    voice=await channel.connect()
                await voice.disconnect()
                if voice and voice.is_connected():
                    await voice.move_to(channel)
                else:
                    voice=await channel.connect()
                await makeembed('음성채널 입장 완료', '응용 기능')
            elif message.content=='키키야 퇴장':
                channel=message.author.voice.channel
                voice=get(client.voice_clients, guild=message.guild)
                if voice and voice.is_connected():
                    await voice.disconnect()
                    await makeembed('음성채널 퇴장 완료', '응용 기능')
                else:
                    await makeembed('`키키야 입장` 안하나여', '에러 감지 기능')
            elif message.content.startswith('키키야 재생'):
                def check_queue():
                    Queue_infile=os.path.isdir('./Queue')
                    if Queue_infile is True:
                        DIR=os.path.abspath(os.path.realpath('Queue'))
                        length=len(os.listdir(DIR))
                        try:
                            first_file=os.listdir(DIR)[0]
                        except:
                            queues.clear()
                            return 
                        main_location=os.path.dirname(os.path.realpath(__file__))
                        song_path=os.path.abspath(os.path.realpath('Queue')+'\\'+first_file)
                        if length!=0:
                            song_there=os.path.isfile('song.mp3')
                            if song_there:
                                os.remove('song.mp3')
                            shutil.move(song_path, main_location)
                            for file in os.listdir('./'):
                                if file.endswith('.mp3'):
                                    os.rename(file, 'song.mp3')
                            voice.play(discord.FFmpegPCMAudio('song.mp3'), after=lambda e: check_queue()    )
                            voice.source=discord.PCMVolumeTransformer(voice.source)
                            voice.source.volume=0.07
                        else:
                            queues.clear()
                            return
                    else:
                        queues.clear()
                url=message.content[7:]
                song_there=os.path.isfile('song.mp3')
                try:
                    if song_there:
                        os.remove('song.mp3')
                        queues.clear()
                except PermissionError:
                    await makeembed('음악 재생중입니다', '에러 감지 기능')
                    return
                Queue_infile=os.path.isdir('./Queue')
                try:
                    Queue_folder='./Queue'
                    if Queue_infile is True:
                        shutil.rmtree(Queue_folder)
                except:
                    print('')
                await makeembed('준비됨', '응용 기능')
                voice=get(client.voice_clients, guild=message.guild)
                ydl_ops={
                    'format':'bestaudio/best',
                    'default_search':'ytsearch',
                    'postprocessors':[{
                        'key':'FFmpegExtractAudio',
                        'preferredcodec':'mp3',
                        'preferredquality':'192',
                    }],
                }
                with youtube_dl.YoutubeDL(ydl_ops) as ydl:
                    ydl.download([url])
                for file in os.listdir('./'):
                    if file.endswith('.mp3'):
                        name=file
                        os.rename(file, 'song.mp3')
                voice.play(discord.FFmpegPCMAudio('song.mp3'), after=lambda e: check_queue())
                voice.source=discord.PCMVolumeTransformer(voice.source)
                voice.source.volume=0.20
                nname=name.rsplit('-', 2)
                await makeembed(f'{nname} 재생중', '응용 기능')
            elif message.content=='키키야 일시정지':
                voice=get(client.voice_clients, guild=message.guild)
                if voice and voice.is_playing():
                    voice.pause()
                    await makeembed('일시정지 완료', '응용 기능')
                else:
                    await makeembed('음악 재생 하고 말해!', '에러 감지 기능')
            elif message.content=='키키야 다시재생':
                voice=get(client.voice_clients, guild=message.guild)
                if voice and voice.is_paused():
                    await makeembed('다시재생 완료', '응용 기능')
                    voice.resume()
                else:
                    await makeembed('`키키야 일시정지`를 하고 하세여!', '에러 감지 기능')
            elif message.content=='키키야 스킵':
                voice=get(client.voice_clients, guild=message.guild)
                queues.clear()
                if voice and voice.is_playing():
                    await makeembed('스킵 완료', '응용 기능')
                    voice.stop()
                else:
                    await makeembed('음악이나 듣고 말하세여', '에러 감지 기능')
            elif message.content.startswith('키키야 예약'):
                url=message.content[7:]
                Queue_infile=os.path.isdir("./Queue")
                if Queue_infile is False:
                    os.mkdir("Queue")
                DIR=os.path.abspath(os.path.realpath("Queue"))  
                q_num=len(os.listdir(DIR))
                q_num+=1
                add_queue=True 
                while add_queue:
                    if q_num in queues:
                        q_num+=1
                    else:
                        add_queue=False
                        queues[q_num]=q_num
                queue_path=os.path.abspath(os.path.realpath("Queue")+f'/song{q_num}.%(ext)s')
                ydl_ops={
                    'format':'bestaudio/best',
                    'default_search':'ytsearch',
                    'quite':True,
                    'outtmpl':queue_path,
                    'postprocessors':[{
                        'key':'FFmpegExtractAudio',
                        'preferredcodec':'mp3',
                        'preferredquality':'192',
                    }],
                }
                with youtube_dl.YoutubeDL(ydl_ops) as ydl:
                    ydl.download([url])
                await makeembed('재생목록에 음악이 추가되었습니다', '응용 기능')
            elif message.content.startswith('키키야 엑셀편집'):
                if str(message.author.id)!=str(admin):
                    await message.channel.send('이런 명령어는 도데체 언제 알았데','관리자 기능')
                    return 
                strr=message.content[9:]
                cell=strr.split('/')
                file=openpyxl.load_workbook(cell[0])
                sheet=file.active
                sheet[cell[1]].value=str(cell[2])
                file.save(cell[0])
                await makeembed('엑셀파일 편집 성공입니다', '관리자 기능')
            elif message.content.startswith('키키야 위키'):
                wiki=wikipediaapi.Wikipedia(language='ko', extract_format=wikipediaapi.ExtractFormat.WIKI)
                keyword=message.content[7:]
                if str(wiki.page(keyword).exists())!=str('True'):
                    await makeembed('없는 문서입니다', '에러 감지 기능')
                    return
                await makeembed('위키피디아 내용을 디엠으로 전송했습니다(2000자까지만 보내집니다)', '응용 기능')
                now=datetime.datetime.now()
                embed=discord.Embed(title=f'위키피디아-{keyword}', description=wiki.page(keyword).text[0:2000], colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                await message.author.send(embed=embed)
            elif message.content=='키키야 돈줘':
                asdfa=cur.execute('SELECT * FROM USERS')
                for row in asdfa:
                    if str(message.author.id) in row:
                        await makeembed('1원부터 1000원 까지 아무 양의 돈을 줍니다', '돈 기능')
                        if str(message.author.id) in vip:
                            await makeembed('1원부터 10000원 까지 아무 양의 돈을 줍니다', '돈 기능')
                            rand=random.randrange(1, 10001)
                        else:
                            await makeembed('1원부터 1000원 까지 아무 양의 돈을 줍니다', '돈 기능')
                            rand=random.randrange(1, 1001)
                        cur.execute(f'SELECT money FROM USERS WHERE id={str(message.author.id)}')
                        money=int(cur.fetchall()[0][0])+rand
                        cur.execute(f'UPDATE USERS SET money={money} WHERE id={str(message.author.id)}')
                        await makeembed(f'**{rand}**원을 줬습니다\n현재 **{message.author.display_name}**님의 돈 액수: {money}', '돈 기능')
                        con.commit()
                        return
                await makeembed("`키키야 가입`을 입력해 먼저 가입을 하세요", '에러 감지 기능')
            elif message.content.startswith('키키야 배팅'):
                money=int(message.content[7:])
                asdfa=cur.execute('SELECT * FROM USERS')
                for row in asdfa:
                    if str(message.author.id) in row:
                        if row[2]<int(money):
                            await makeembed('돈이나 벌고 오세여', '돈 기능')
                            return  
                        await makeembed('성공하면 그 돈의 2배를 얻지만 실패하면 그 돈의 양만큼 뺏깁니다', '돈 기능')
                        if str(message.author.id) in vip:
                            rand=random.randrange(1, 4)
                        else:
                            rand=random.randrange(1, 3)
                        if rand==1 or rand==3:
                            cur.execute(f'UPDATE USERS SET money={row[2]+money*2} WHERE id={str(message.author.id)}')
                            await makeembed('성공하셨습니다!!\n~~역시 도박은 몸에 좋아여~~', '돈 기능')
                        elif rand==2:
                            cur.execute(f'UPDATE USERS SET money={row[2]-money} WHERE id={str(message.author.id)}')
                            await makeembed('실패하셨습니다!!\n~~역시 도박은 몸에 안좋아여~~', '돈 기능')
                        con.commit()
                        return
                await makeembed("`키키야 가입`을 입력해 먼저 가입을 하세요", '에러 감지 기능')
            elif message.content.startswith('키키야 추방'):
                if str(message.author.id)!=str(admin):
                    await makeembed('추방해서 뭐하게요', '관리자 기능')
                    return None
                user=message.guild.get_member(int(message.content[7:]))
                await message.guild.kick(user)
                await makeembed('유저 추방 완료', '관리자 기능')
            elif message.content.startswith('키키야 경고'):
                user=message.content[7:]
                user=message.guild.get_member(int(user))
                user=str(user.id)
                asdf=cur.execute('SELECT * FROM USERS')
                for row in asdf:
                    if user in row:
                        warn=row[6]+1
                        cur.execute(f'UPDATE USERS SET warnings={warn} WHERE ID={user}')
                        if warn>=7:
                            cur.execute(f'UPDATE USERS SET black=1 WHERE ID={user}')
                            await makeembed('경고를 1회주고 경고가 7회 넘은 유저여서 블랙리스트에 집어(?) 넣었습니다', '관리자 기능')
                            con.commit()
                            return None
                        else:
                            await makeembed('경고를 1회 줬습니다', '관리자 기능')
                            con.commit()
                            return None
                await makeembed('가입이 안된 사용자입니다', '에러 감지 기능')
            elif message.content=='키키야 올인':
                asdfa=cur.execute('SELECT * FROM USERS')
                for row in asdfa:
                    if str(message.author.id) in row:
                        if str(message.author.id) in vip:
                            await makeembed('성공하면 현재 돈의 2배를 주지만 실패하면 돈은 0원이 됩니다\n확률은 3/2!', '돈 기능')
                            rand=random.randrange(1, 4)
                        else:
                            await makeembed('성공하면 현재 돈의 2배를 주지만 실패하면 돈은 0원이 됩니다\n확률은 2/1!', '돈 기능')
                            rand=random.randrange(1, 3)
                        if rand==1 or rand==3:
                            cur.execute(f'UPDATE USERS SET money={row[2]*3} WHERE id={str(message.author.id)}')
                            await makeembed('성공입니다!!!', '돈 기능')
                        elif rand==2:
                            cur.execute(f'UPDATE USERS SET money=0 WHERE id={str(message.author.id)}')
                            await makeembed('실패입니다!!!(복사 붙여넣기를 해 어쩔수 없이 느낌표를 달아야하는 키키봇의 슬픈 이야기)', '돈 기능')
                        con.commit()
                        return
                await makeembed("`키키야 가입`을 입력해 먼저 가입을 하세요", '에러 감지 기능')
            elif message.content.startswith("키키야 롤"):
                Name=message.content.split(' ')[2]
                SummonerName = "" 
                TierUnranked = "" 
                LeagueType = [] 
                Tier = [] 
                LP = [] 
                Wins = [] 
                Losses = [] 
                Ratio = []
                url='https://www.op.gg/summoner/userName=' + Name 
                hdr = {'Accept-Language': 'ko_KR,en;q=0.8', 'User-Agent': ('Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.70 Mobile Safari/537.36')} 
                req = requests.get(url, headers=hdr) 
                html = req.text 
                soup = BeautifulSoup(html, 'html.parser')
                for i in soup.select('div[class=SummonerName]'): 
                    SummonerName = i.text 
                TierUnranked = soup.select('div.Cell')
                for i in soup.select('div[class=LeagueType]'): 
                    LeagueType.append(i.text) 
                for i in soup.select('div[class=Tier]'): 
                    Tier.append(i.text)
                for i in soup.select('div[class=LP]'): 
                    LP.append(i.text) 
                for i in soup.select('span[class=Wins]'): 
                    Wins.append(i.text) 
                for i in soup.select('span[class=Losses]'): 
                    Losses.append(i.text) 
                for i in soup.select('span[class=Ratio]'): 
                    Ratio.append(i.text)
                now=datetime.datetime.now()
                embed = discord.Embed(
                    title='롤 정보',
                    description='응용 기능',
                    colour=discord.Colour.blue()
                ).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
                if SummonerName != "":
                    if 'Unranked' in str(TierUnranked[0]): 
                        embed.add_field(name='솔로랭크', value='Unranked', inline=False)
                    else: 
                        asdf=Tier[0].strip('\n\t')
                        embed.add_field(name='솔로랭크', value=f'티어: {asdf}\nLP: {LP[0]}\n승/패: {Wins[0]}/{Losses[0]}\n승률: {Ratio[0]}', inline=False)
                    if 'Unranked' in str(TierUnranked[1]): 
                        embed.add_field(name='자유랭크', value='Unranked', inline=False)
                    else: 
                        asdf=Tier[1].strip('\n\t')
                        embed.add_field(name='자유랭크', value=f'티어: {asdf}\nLP: {LP[1]}\n승/패: {Wins[1]}/{Losses[1]}\n승률: {Ratio[1]}', inline=False)
                else:
                    embed.add_field(name="에러", value="사용자 정보 없음")
                await message.channel.send( embed=embed)
            elif message.content.startswith('키키야 마크서버'):
                server=message.content.split(' ')[2]
                url='https://mcsrvstat.us/server/'+str(server)
                res=requests.get(   url)
                soup=BeautifulSoup(res.content, 'html.parser')
                clas=soup.findAll('h2')
                try:
                    adsf=clas[1]
                    if adsf==None:
                        return
                except IndexError:
                    await makeembed('서버가 온라인입니다', '응용 기능')
                    return
                await makeembed('서버가 오프라인입니다', '응용 기능')
            elif message.content=='키키야 재부팅':
                if str(message.author.id)!=str(admin):
                    await makeembed('재부팅을 왜하게요', '관리자 기능')
                    return      
                await makeembed('재부팅 되었습니다', '관리자 기능')
                await client.close()
                await client.logout()
            elif message.content=='키키야 내돈':
                asdfa=cur.execute('SELECT * FROM USERS')
                for row in asdfa:
                    if str(message.author.id) in row:
                        await makeembed(f'{message.author.display_name}님의 돈: {str(row[2])}원', '돈 기능')
                        return
                await makeembed("`키키야 가입`을 입력해 먼저 가입을 하세요", '에러 감지 기능')
            elif message.content=='키키야 주사위':
                now=datetime.datetime.now()
                msg=await message.channel.send(embed=discord.Embed(title='과연...', descirption='응용 기능', colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일'))
                await asyncio.sleep(3)
                await msg.edit(embed=discord.Embed(title=random.randint(1, 6), description='응용 기능', colour=discord.Colour.blue()).set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일'))
            else:
                learn=["kiki"]
                q=message.content[4:]
                flag=0
                asdf=cur.execute('SELECT * FROM CUSTOMCOMMANDS')
                for row in asdf:
                    if row[0] == q:
                        learn.append(str(row[1]))
                        flag+=1
                if flag!=0: 
                    if len(learn)>2:
                        rand=random.randrange(1, len(learn)+1)
                        await makeembed(f'{learn[rand]}', '챗 기능')
                    else:
                        await makeembed(f'{learn[1]}', '챗 기능')
                else: 
                    await makeembed('키키봇이 모르는 명령어', '키키봇이 모르는 명령어')
    except ZeroDivisionError:
        await makeembed('0으로 나누기가 가능하나...', '에러 감지 기능')
    except discord.errors.Forbidden:
        await makeembed('권한 주떼여~', '에러 감지 기능')
    except IndexError:
        await makeembed('에휴, 입력 형식을 안지켰구먼', '에러 감지 기능')
    except SyntaxError:
        await makeembed('키키야, 제발 오타좀 치지 마라...', '에러 감지 기능')
    #except Exception as exc:
    #    now=datetime.datetime.now()
    #    embed=discord.Embed(
    #        title=f'에러 내용: {str(exc)}',
    #        description='에러 감지 기능',
    #        colour=discord.Colour.blue()
    #    ) 
    #    embed.set_footer(icon_url=message.author.avatar_url, text=f' | {str(message.author.display_name)} | {str(now.year)}년 {str(now.month)}월 {str(now.day)}일')
    #    await message.channel.send(embed=embed)
async def my_background_task():
    await client.wait_until_ready()
    while not client.is_closed():
        act=["키키야 도움말을 입력해 명령어 확인", f'{len(client.guilds)}개의 서버에 참여중', f'{len(client.users)}명의 유저들과 소통하는중', '문의는 키키#1778', '키키야 서버 를 이용해 키키봇 서포트로']
        for i in act:
            game = discord.Game(str(i))
            await client.change_presence(status=discord.Status.online  , activity=game)
            await asyncio.sleep(5)
client.loop.create_task(my_background_task())
client.run(token)
os.execv(sys.executable, ['python']+sys.argv)
