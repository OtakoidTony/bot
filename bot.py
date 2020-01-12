import asyncio
import discord
from discord.utils import get
import random
import threading
import time
import bs4
import logging
import openpyxl 
from discord.ext import tasks
from itertools import cycle
import youtube_dl
import re
import os
from urllib.request import urlopen, Request
import urllib
import urllib.request
import json
app=discord.Client()
token = "This is a secret"
@app.event
async def on_ready():
    print("다음으로 로그인합니다 : ")
    print(app.user.name)
    print(app.user.id)
    print("==========")
@app.event
async def on_member_join(member):
    channel = app.get_channel(660775449449988107)
    await channel.send('{}, 환영합니다^^'.format(member.mention))
@app.event
async def on_member_remove(member):
    channel = app.get_channel(660775449449988107)
    await channel.send('{}, 가지마요ㅠㅠ'.format(member.mention))
@app.event
async def on_message(message):
    try:
        if message.content.startswith('키키야'):
            if message.author.bot:
                 return None
            if message.content == "키키야": await message.channel.send("안녕하세요 키키봇입니다^^")
            elif message.content == "키키야 테스트": await message.channel.send(message.author)
            elif message.content == '키키야 욕': await message.channel.send('~~**||__*삐릴리 삐약삐약 심한욕*__||**~~')
            elif message.content == "키키야 맨션": await message.channel.send("오타겠죠... 정확한건 멘션인데")
            elif message.content == '키키야 멘션': await message.channel.send('<@{}>, 멘션 여기요^^'.format(message.author.id))
            elif message.content == "키키야 일본": await message.channel.send(":flag_jp::left_facing_fist:")
            elif message.content == "키키야 한국":  await message.channel.send(":flag_kr::thumbsup:")   
            elif message.content == "키키야 대한민국": await message.channel.send( ":flag_kr::thumbsup:")
            elif message.content == "키키야 북한": await message.channel.send( ":flag_kp:")
            elif message.content == '키키야 뒤져': await message.channel.send( '그럼 님은 앞져요(아재키키)')
            elif message.content == '키키야 바보': await message.channel.send( '제가 바다의 보물이죠(수중생활)')
            elif message.content == '키키야 천재': await message.channel.send( '저는 천하의 재수있는놈(?) 입니다')
            elif message.content == '키키야 메롱': await message.channel.send( '메롱x무한')
            elif message.content == '키키야 아무말': await message.channel.send( '아니, 아무말 말고 다른 말을 써야져')
            elif message.content == '키키야 안녕': await message.channel.send( '안녕하세요')
            elif message.content == '키키야 키키는': await message.channel.send( '이 봇을 만드셨죠')
            elif message.content == '키키야 키키': await message.channel.send( '이 봇을 만드셨죠')
            elif message.content == '키키야 디토': await message.channel.send( '엄청난 슷칼러(?)에요')
            elif message.content == '키키야 디토는': await message.channel.send( '엄청난 슷칼러(?)에요')
            elif message.content == '키키야 꿀벌': await message.channel.send( '꿀벌꿀벌~')
            elif message.content == '키키야 내아바타': await message.channel.send('멋진 아바타네요^^ {}'.format(message.author.avatar_url))
            elif message.content == '키키야 꿀벌은': await message.channel.send( '꿀벌꿀벌~')
            elif message.content == '키키야 내프사': await message.channel.send('멋진 프사네요^^ {}'.format(message.author.avatar_url))
            elif message.content == '키키야 초대': await message.channel.send('키키봇을 초대하고싶으면 `키키야 봇초대`, 키키봇 서포트 서버에 초대받고 싶으면 `키키야 서버초대`')
            elif message.content == '키키야 mswgen':  await message.channel.send('엄청난 바보(천재)죠')
            elif message.content == '키키야 mswgen은': await message.channel.send( '엄청난 바보(천재)죠')
            elif message.content == '키키야 미안해':  await message.channel.send('미안해미안해하지마')
            elif message.content == '키키야 도움': await message.channel.send( 'ㄴㄴ 도움말')
            elif message.content == '키키야 생일':  await message.channel.send('몸은 12월 11일에 만들어졌지만 뇌는 디토봇과 같은 12월 22일에 만들어졌어요^^')
            elif message.content == '키키야 멍청이': await message.channel.send( '저 멍청이로 삼행시 지으면 님 멍청이에요')
            elif message.content == '키키야 멍청이삼행시': await message.channel.send( '멍: 멍청한 님\n청: (포기)\n이: 이름은?\n(키키봇 인성논란)')
            elif message.content == '키키야 죽어': await message.channel.send( '형님 먼저~')
            elif message.content == '키키야 인성':await message.channel.send(  '아, 참 생각해볼 문제군')
            elif message.content == '키키야 ㅎ2': await message.channel.send( 'ㅎ3')
            elif message.content == '키키야 애교': await message.channel.send( '1+1은? 귀욮ㅍ:left_facing_fist::left_facing_fist::left_facing_fist:\n토나오겠다')
            elif message.content == '키키야 팀크레센도': await message.channel.send( '아 저기 안에 들어갈수만 있다면...')
            elif message.content == '키키야 슷칼봇': await message.channel.send( '저안테는 워렌 버펫같은 대단한 투자가이자 존경하는 분이죠')
            elif message.content == '키키야 배추봇': await message.channel.send( '아 인성 안좋은 선배죠')
            elif message.content == '키키야 배그': await message.channel.send( '아 그 설치 오래걸리는 게임')
            elif message.content == '키키야 클로': await message.channel.send( '제가 그거 1달만에 아레나 3에서 10까지 간 사연 아나요?')
            elif message.content == '키키야     클오클': await message.channel.send( '아 그거 저 하다가 탈탈 털림요')
            elif message.content == '키키야 언더테일':  await message.channel.send('그 해골나오는거요?')
            elif message.content == '키키야 행복한서버는': await message.channel.send( '제 집이죠')
            elif message.content=='키키야 도배': await message.channel.send('키키야 도배')
            elif message.content=='키키야 랜덤게임': await message.channel.send('키키야 랜덤을 입력해 번호가 더 큰 사람이 이겨요^^')
            elif message.content=='키키야 랜덤': await message.channel.send('<@{}>, '.format(message.author.id)+str(random.randrange(1, 1001)))
            elif message.content=='키키야 개논리': await message.channel.send('저의 선배님인 헤이창봇형 따라하다 헤이창봇형안테 혼났어요ㅠㅠ(그냥 없어졌다는 얘기)')
            elif message.content == '키키야 상상도못한정체': await message.channel.send('상상도 못한 정체https://giphy.com/gifs/f67Fu9MlQz0MxikSVe')
            elif message.content == '키키야 서버초대': await message.channel.send( 'https://discord.gg/m2zNAS3')
            elif message.content == '키키야 봇초대': await message.channel.send( 'https://discordapp.com/api/oauth2/authorize?client_id=657949896263073811&permissions=8&scope=bot')
            elif message.content=='키키야 심심해':
                r=random.randrange(1, 5)
                if r==1:
                    embed=discord.Embed(
                        title= '개논리',
                        description= '저는 개입니다\n제가 논리를 펼쳤습니다\n그래서 그 논리는 개논리입니다',
                        footer= '제작자: 키키',
                        colour= discord.Colour.blue()
                    )
                    await message.channel.send(embed=embed)
                elif r==2:
                    await message.channel.send('꽃집 주인이 가장 싫어하는 도시는?')
                    await message.channel.send('||시드니||')
                elif r==3:
                    await message.channel.send('```키키야 가위바위보 (바위, 보자기, 가위 중 하나)```를 입력해 키키와 가위바위보를 하세여')
                elif r==4:
                    await message.channel.send('```키키야 랜덤```를 입력해 랜덤게임을 즐기세여^^')
            elif message.content=='키키야 주사위':
                await message.channel.send('주사위 값: '+str(random.randrange(1, 7)))
            elif message.content.startswith('키키야 실시간검색어') or message.content.startswith('키키야 실검'):
                url = "https://www.naver.com/"
                html = urllib.request.urlopen(url)
                bsObj = bs4.BeautifulSoup(html, "html.parser")
                realTimeSerach1 = bsObj.find('div', {'class': 'ah_roll_area PM_CL_realtimeKeyword_rolling'})
                realTimeSerach2 = realTimeSerach1.find('ul', {'class': 'ah_l'})
                realTimeSerach3 = realTimeSerach2.find_all('li')
                embed = discord.Embed(
                    title='네이버 실시간 검색어',
                    description='실시간검색어',
                    colour=discord.Colour.green()
                )
                for i in range(0,20):
                    realTimeSerach4 = realTimeSerach3[i]
                    realTimeSerach5 = realTimeSerach4.find('span', {'class': 'ah_k'})
                    realTimeSerach = realTimeSerach5.text.replace(' ', '')
                    realURL = 'https://search.naver.com/search.naver?ie=utf8&query='+realTimeSerach
                    embed.add_field(name=str(i+1)+'위', value='\n'+'[%s](<%s>)' % (realTimeSerach, realURL), inline=False) # [텍스트](<링크>) 형식으로 적으면 텍스트 하이퍼링크 만들어집니다
                await message.channel.send( embed=embed)
            elif message.content.startswith("키키야 롤"):
                learn = message.content.split(" ")
                location = learn[2]
                enc_location = urllib.parse.quote(location)
                url = "http://www.op.gg/summoner/userName=" + enc_location
                html = urllib.request.urlopen(url)
                bsObj = bs4.BeautifulSoup(html, "html.parser")
                rank1 = bsObj.find("div", {"class": "TierRankInfo"})
                rank2 = rank1.find("div", {"class": "TierRank"})
                rank3 = rank2.find("span", {"class": "tierRank"})
                rank4 = rank3.text  # 티어표시 (브론즈1,2,3,4,5 등등)
                if rank4 != 'Unranked':
                    jumsu1 = rank1.find("div", {"class": "TierInfo"})
                    jumsu2 = jumsu1.find("span", {"class": "LeaguePoints"})
                    jumsu3 = jumsu2.text
                    jumsu4 = jumsu3.strip()#점수표시 (11LP등등)
                    winlose1 = jumsu1.find("span", {"class": "WinLose"})
                    winlose2 = winlose1.find("span", {"class": "wins"})
                    winlose2_1 = winlose1.find("span", {"class": "losses"})
                    winlose2_2 = winlose1.find("span", {"class": "winratio"})
                    winlose2txt = winlose2.text
                    winlose2_1txt = winlose2_1.text
                    winlose2_2txt = winlose2_2.text 
                    embed = discord.Embed(
                        title='롤 정보',
                        description='롤 정보입니다.',
                        colour=discord.Colour.green()
                    )
                if rank4=='Unranked':
                    embed.add_field(name='당신의 티어', value=rank4, inline=False)
                    embed.add_field(name='-당신은 언랭-', value="언랭은 더이상의 정보는 제공하지 않습니다.", inline=False)
                    await message.channel.send( embed=embed)
                else:
                    embed.add_field(name='당신의 티어', value=rank4, inline=False)
                    embed.add_field(name='당신의 LP(점수)', value=jumsu4, inline=False)
                    embed.add_field(name='당신의 승,패 정보', value=winlose2txt+" "+winlose2_1txt, inline=False)
                    embed.add_field(name='당신의 승률', value=winlose2_2txt, inline=False)
                    await message.channel.send( embed=embed)
            elif message.content.startswith('키키야 뮤트'):
                author=message.guild.get_member(int(message.content[7:]))
                role=discord.utils.get(message.guild.roles, name='뮤트')
                await author.add_roles(role)
                await message.channel.send('뮤트 성공')
            elif message.content.startswith('키키야 언뮤트'):
                author=message.guild.get_member(int(message.content[8:]))
                role=discord.utils.get(message.guild.roles, name='뮤트')
                await author.remove_roles(role)
                await message.channel.send('언뮤트 성공')
            elif message.content.startswith('키키야 역할줘'):
                author=message.guild.get_member(int(message.content[8:26]))
                role=discord.utils.get(message.guild.roles, name=message.content[28:])
                await author.add_roles(role)
                await message.channel.send('성공적으로 역할을 줬습니다')
            elif message.content.startswith("키키야 날씨"):
                learn = message.content.split(" ")
                location = learn[2]
                enc_location = urllib.parse.quote(location+'날씨')
                hdr = {'User-Agent': 'Mozilla/5.0'} 
                url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + enc_location
                req = Request(url, headers=hdr)
                html = urllib.request.urlopen(req)
                bsObj = bs4.BeautifulSoup(html, "html.parser")
                todayBase = bsObj.find('div', {'class': 'main_info'})
                todayTemp1 = todayBase.find('span', {'class': 'todaytemp'})
                todayTemp = todayTemp1.text.strip()  # 온도
                todayValueBase = todayBase.find('ul', {'class': 'info_list'})
                todayValue2 = todayValueBase.find('p', {'class': 'cast_txt'})
                todayValue = todayValue2.text.strip()  
                todayFeelingTemp1 = todayValueBase.find('span', {'class': 'sensible'})
                todayFeelingTemp = todayFeelingTemp1.text.strip()  
                todayMiseaMongi1 = bsObj.find('div', {'class': 'sub_info'})
                todayMiseaMongi2 = todayMiseaMongi1.find('div', {'class': 'detail_box'})
                todayMiseaMongi3 = todayMiseaMongi2.find('dd')
                todayMiseaMongi = todayMiseaMongi3.text  
                tomorrowBase = bsObj.find('div', {'class': 'table_info weekly _weeklyWeather'})
                tomorrowTemp1 = tomorrowBase.find('li', {'class': 'date_info'})
                tomorrowTemp2 = tomorrowTemp1.find('dl')
                tomorrowTemp3 = tomorrowTemp2.find('dd')
                tomorrowTemp = tomorrowTemp3.text.strip()  
                tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
                tomorrowMoring1 = tomorrowAreaBase.find('div', {'class': 'main_info morning_box'})
                tomorrowMoring2 = tomorrowMoring1.find('span', {'class': 'todaytemp'})
                tomorrowMoring = tomorrowMoring2.text.strip()  
                tomorrowValue1 = tomorrowMoring1.find('div', {'class': 'info_data'})
                tomorrowValue = tomorrowValue1.text.strip()  
                tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
                tomorrowAllFind = tomorrowAreaBase.find_all('div', {'class': 'main_info morning_box'})
                tomorrowAfter1 = tomorrowAllFind[1]
                tomorrowAfter2 = tomorrowAfter1.find('p', {'class': 'info_temperature'})
                tomorrowAfter3 = tomorrowAfter2.find('span', {'class': 'todaytemp'})
                tomorrowAfterTemp = tomorrowAfter3.text.strip() 
                tomorrowAfterValue1 = tomorrowAfter1.find('div', {'class': 'info_data'})
                tomorrowAfterValue = tomorrowAfterValue1.text.strip()
                embed = discord.Embed(
                    title=learn[1]+ ' 날씨 정보',
                    description=learn[1]+ '날씨 정보입니다.',
                    colour=discord.Colour.gold()
                )
                embed.add_field(name='현재온도', value=todayTemp+'˚', inline=False)  # 현재온도
                embed.add_field(name='체감온도', value=todayFeelingTemp, inline=False)  # 체감온도
                embed.add_field(name='현재상태', value=todayValue, inline=False)  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
                embed.add_field(name='현재 미세먼지 상태', value=todayMiseaMongi, inline=False)  # 오늘 미세먼지
                embed.add_field(name='오늘 오전/오후 날씨', value=tomorrowTemp, inline=False)  # 오늘날씨 # color=discord.Color.blue()
                embed.add_field(name='**----------------------------------**',value='**----------------------------------**', inline=False)  # 구분선
                embed.add_field(name='내일 오전온도', value=tomorrowMoring+'˚', inline=False)  # 내일오전날씨
                embed.add_field(name='내일 오전날씨상태, 미세먼지 상태', value=tomorrowValue, inline=False)  # 내일오전 날씨상태
                embed.add_field(name='내일 오후온도', value=tomorrowAfterTemp + '˚', inline=False)  # 내일오후날씨
                embed.add_field(name='내일 오후날씨상태, 미세먼지 상태', value=tomorrowAfterValue, inline=False)  # 내일오후 날씨상태
                await message.channel.send(embed=embed)
            elif message.content=='키키야 봇정보':
                embed=discord.Embed(
                    title="키키봇 정보",
                    description=f"현재 {len(app.guilds)}개의 서버에 참여중, {len(app.users)}명의 유저들과 소통중\n\n**키키봇이 참여하고 있는 서버**"
                )
                for i in app.guilds:
                    embed.add_field(name=i, value=i)
                await message.channel.send(embed=embed)
            elif message.content=='키키야 공격':
                file=openpyxl.load_workbook('xp.xlsx')
                sheet=file.active
                if sheet['A'+str(1)].value==None:
                    sheet['A'+str(1)].value=0
                sheet['A'+str(1)].value+=1
                await message.channel.send('`현재 키키봇 xp: {}`'.format(str(sheet['A'+str(1)].value)))
                file.save('xp.xlsx')
            elif message.content=='키키야 핑':
                embed = discord.Embed(
                    title = 'PONG!',
                    author = '키키봇',
                    colour = discord.Colour.blue()
                )
                embed.add_field(name = 'Result', value = f' {round(app.latency * 1000)}ms' + f' (0.{round(app.latency * 1000)}초)', inline = False)
                await message.channel.send(embed = embed)
            elif message.content.startswith('키키야 가위바위보'):
                ohyeah=message.content.split(' ')
                value=ohyeah[2]
                mine=random.randrange(1, 4) #1은 바위, 2는 가위, 3은 보자기
                if value=='가위':
                    if mine==3:
                        await message.channel.send('님이 승리!!!')
                    elif mine==2:
                        await message.channel.send('무승부')
                    else:
                        await message.channel.send('님이 패베')
                elif value=='바위':
                    if mine==2:
                        await message.channel.send('님이 승리!!!')
                    elif mine==1:
                        await message.channel.send('무승부')
                    else:
                        await message.channel.send('님이 패베')
                elif value=='보자기':
                    if mine==1:
                        await message.channel.send('님이 승리!!!')
                    elif mine==3:
                        await message.channel.send('무승부')
                    else:
                        await message.channel.send('님이 패베')
                else:
                    await message.channel.send('아니 가위, 바위, 보자기 중 내라고!!!\n보자보자 하니까 보자기가 생각나네')
                if mine == 1:
                    await message.channel.send('저는 바위를 냈어요')
                elif mine == 2:
                    await message.channel.send('저는 가위를 냈어요')
                else:
                    await message.channel.send('저는 보자기를 냈어요')
            elif message.content.startswith('키키야 명령어추가'):
                file = openpyxl.load_workbook("기억.xlsx") #파일 이름은 상관 없어요
                sheet = file.active
                strsss=message.content.split('명령어추가')[1]
                q=strsss.split("/")[0] #A부분
                a=strsss.split("/")[1] #B부분
                i = 1
                while sheet["A" + str(i)].value != None:
                    i+=1
                sheet["A" + str(i)].value = str(q[1:]) #A 저장
                sheet["B" + str(i)].value = str(a)     #B 저장
                sheet["C" + str(i)].value = str(message.author.id) #이 말을 가르쳐준 사람 id 저장
                sheet["D" + str(i)].value = str(message.author) #이 말을 가르쳐준 사람 저장
                await message.channel.send("[" + str(q[1:]) + "]라고 말하면 [" + str(a) + "]라고 대답하는 것을 배웠어!") #출력
                file.save("기억.xlsx")
            elif message.content.startswith('키키야 청소'):
                varrr=message.content.split(' ')
                await message.channel.purge(limit=int(varrr[2])+1)
            elif message.content.startswith('키키야 찬반투표'):
                lern=message.content[9:]
                embed=discord.Embed(
                    title=str(lern),
                    description=f"{message.author.display_name}님의 찬반투표"
                )
                msg=await message.channel.send(embed=embed)
                await msg.add_reaction('👍')
                await msg.add_reaction('👎')
            elif message.content.startswith('키키야 투표'):
                strr=message.content[7:]
                strrr=strr.split('/')[0]
                doit=strr.split('/')
                i=1
                embed=discord.Embed(
                    title=doit[0],
                    description=f'주최자: {message.author.display_name}',
                    colour=discord.Colour.blue()
                )
                while i>=len(doit)==False:
                    embed.add_field(name=f'{str(i+1)}번째 항목', value=str(doit[i-1]))
                    i+=1
                msg=await message.channel.send(embed=embed)
                i=1
                while i>=len(doit)==False:
                    if i==1:
                        await msg.add_reaction('1️⃣')
                    elif i==2:
                        await msg.add_reaction('2️⃣')
                    elif i==3:
                        await msg.add_reaction('3️⃣')
                    elif i==4:
                        await msg.add_reaction('4️⃣')
                    elif i==5:
                        await msg.add_reaction('5️⃣')
                    elif i==6:
                        await msg.add_reaction('6️⃣')
                    elif i==7:
                        await msg.add_reaction('7️⃣')
                    elif i==8:
                        await msg.add_reaction('8️⃣')
                    elif i==9:
                        await msg.add_reaction('9️⃣')
                    elif i==10:
                        await msg.add_reaction('🔟')
                    i+=1
                await message.channel.send(len(doit))
            elif message.content.startswith('키키야 내정보'):
                roles=[role for role in message.author.roles]
                embed=discord.Embed(colour=message.author.color, timestamp=message.created_at)
                embed.set_author(name=f"유저정보 - {message.author}")
                embed.set_thumbnail(url=message.author.avatar_url)
                embed.set_footer(text=f"{message.author}에게 요청받음", icon_url=message.author.avatar_url)
                embed.add_field(name="아이디", value=message.author.id)
                embed.add_field(name="닉네임", value=message.author.display_name)
                embed.add_field(name="계정 생성 시간", value=message.author.created_at.strftime("%a, %#d %B %Y, %I:%M %p UTC"))
                embed.add_field(name="가입 시간", value=message.author.joined_at.strftime("%a, %#d %B %Y, %I:%M %p UTC"))
                embed.add_field(name=f"가진 역할들 ({len(roles)})", value=" ".join([role.mention for role in roles]))
                embed.add_field(name="가장 높은 역할", value=message.author.top_role.mention)
                embed.add_field(name="상태", value=message.author.status)
                embed.add_field(name="봇", value=message.author.bot)
                await message.channel.send(embed=embed)
            elif message.content=='키키야 도움말':
                embed=discord.Embed(
                    title='키키봇 도움말', 
                    description='키키봇의 명령어들 입니다',
                    colour=discord.Colour.blue()
                )
                embed.add_field(name = '키키야 도움말', value = '도움말')
                embed.add_field(name = '키키야 명령어추가 [입력]/[출력]', value = '창님이 알려줌요')
                embed.add_field(name = '키키야 주사위', value = '주사위 돌리기')
                embed.add_field(name = '키키야 이모지', value = '꒰⑅ᵕ༚ᵕ꒱')
                embed.add_field(name = '키키야 심심해', value = '심심하면 해보세여')
                embed.add_field(name = '키키야 청소 [숫자]', value = '메시지 삭제')
                embed.add_field(name = '키키야 내정보', value = '자기 정보')
                embed.add_field(name = '키키야 공격', value = '키키봇 xp+1')
                embed.add_field(name = '키키야 돈줘', value = '돈받기')
                embed.add_field(name = '키키야 찬반투표 [투표이름]', value = '찬반투표')
                embed.add_field(name = '키키야 봇정보', value = '봇정보')
                embed.add_field(name = '키키야 음성퇴장', value = '유저가 들어가있는 음성채널에 키키봇이 퇴장')
                embed.add_field(name = '키키야 음성입장', value = '유저가 들어가있는 음성채널에 키키봇이 입장')
                embed.add_field(name = '키키야 음성재생 [유튜브 링크]', value = '음악재생')
                embed.add_field(name = '키키야 올인', value = '자기 돈 올인')
                embed.add_field(name = '키키야 디엠전송 [디엠보낼사람 아이디]/[할말]', value = '키키봇이 대신 디엠을 보내줍니다')
                embed.add_field(name = '키키야 말해 [할말]', value = '키키봇이 대신 그 말을 해줍니다')
                embed.add_field(name = '키키야 내돈', value = '자기돈을 확인합니다')
                embed.add_field(name = '키키야 무단침입 [채널아이디] [할말]', value = '키키봇이 그 채널로 무단침입을 합니다')
                embed.add_field(name = '키키야 랜덤게임', value = '랜덤게임 규칙을 알려줘여')
                embed.add_field(name = '키키야 랜덤', value = '1부터 1000까지의 아무 숫자를 알려줘여')
                embed.add_field(name = '키키야 내프사', value = '자기 프사를 보고싶으면 ㄱㄱ')
                embed.add_field(name = '키키야 가위바위보 [가위, 바위, 보자기 중 하나]', value = '누가 이길까여?')
                await message.channel.send(embed=embed)
            elif message.content=='키키야 이모지':
                emoji = [" ꒰⑅ᵕ༚ᵕ꒱ ", " ꒰◍ˊ◡ˋ꒱ ", " ⁽⁽◝꒰ ˙ ꒳ ˙ ꒱◜⁾⁾ ", " ༼ つ ◕_◕ ༽つ ", " ⋌༼ •̀ ⌂ •́ ༽⋋ ",
                    " ( ･ิᴥ･ิ) ", " •ө• ", " ค^•ﻌ•^ค ", " つ╹㉦╹)つ ", " ◕ܫ◕ ", " ᶘ ͡°ᴥ͡°ᶅ ", " ( ؕؔʘ̥̥̥̥ ه ؔؕʘ̥̥̥̥ ) ",
                    " ( •́ ̯•̀ ) ",
                    " •̀.̫•́✧ ", " '͡•_'͡• ", " (΄◞ิ౪◟ิ‵) ", " ˵¯͒ བ¯͒˵ ", " ͡° ͜ʖ ͡° ", " ͡~ ͜ʖ ͡° ", " (づ｡◕‿‿◕｡)づ ",
                    " ´_ゝ` ", " ٩(͡◕_͡◕ ", " ⁄(⁄ ⁄•⁄ω⁄•⁄ ⁄)⁄ ", " ٩(͡ï_͡ï☂ ", " ௐ ", " (´･ʖ̫･`) ", " ε⌯(ง ˙ω˙)ว ",
                    " (っ˘ڡ˘ς) ", "●▅▇█▇▆▅▄▇", "╋╋◀", "︻╦̵̵̿╤──", "ー═┻┳︻▄", "︻╦̵̵͇̿̿̿̿══╤─",
                    " ጿ ኈ ቼ ዽ ጿ ኈ ቼ ዽ ጿ ", "∑◙█▇▆▅▄▃▂", " ♋♉♋ ", " (๑╹ω╹๑) ", " (╯°□°）╯︵ ┻━┻ ",
                    " (///▽///) ", " σ(oдolll) ", " 【o´ﾟ□ﾟ`o】 ", " ＼(^o^)／ ", " (◕‿‿◕｡) ", " ･ᴥ･ ", " ꈍ﹃ꈍ "                                                                          " ˃̣̣̣̣̣̣︿˂̣̣̣̣̣̣ ",
                    " ( ◍•㉦•◍ ) ", " (｡ì_í｡) ", " (╭•̀ﮧ •́╮) ", " ଘ(੭*ˊᵕˋ)੭ ", " ´_ゝ` ", " (~˘▾˘)~ ", "ㄴ(°0°)ㄱ"] 
                randomNum = random.randrange(0, len(emoji)) 
                print(emoji[randomNum])
                await message.channel.send(embed=discord.Embed(description=emoji[randomNum]))
            elif message.content.startswith('키키야 말해'):
                strrr=message.content[7:]
                await message.channel.send(strrr)
            elif message.content.startswith('키키야 디엠전송'):
                strrr=message.content[9:]
                msg=strrr.split('/')[1]
                author=message.guild.get_member(int(strrr.split('/')[0]))
                await author.send(msg)
                await message.channel.send('디엠 전송 완료')
            elif message.content.startswith('키키야 무단침입'):
                strrrrrr=message.content[9:]
                channel=app.get_channel(int(strrrrrr.split('/')[0]))
                await message.channel.purge(limit=1)
                await channel.send(str(strrrrrr.split('/')[1]))
            elif message.content=='키키야 음성입장':
                global voice
                channel=message.author.voice.channel
                voice=get(app.voice_clients, guild=message.guild)
                if voice and voice.is_connected():
                    await voice.move_to(channel)
                else:
                    voice=await channel.connect()
                await voice.disconnect()
                if voice and voice.is_connected():
                    await voice.move_to(channel)
                else:
                    voice=await channel.connect()
                    print(f'----키키봇이 {channel}에 입장했습니다----')
                await message.channel.send(f'{channel} 입장 완료')
            elif message.content=='키키야 음성퇴장':
                channel=message.author.voice.channel
                voice=get(app.voice_clients, guild=message.guild)
                if voice and voice.is_connected():
                    await voice.disconnect()
                    print(f'----키키봇이 {channel}에 퇴장했습니다----')
                    await message.channel.send(f'{channel} 퇴장 완료')
                else:
                    print(f'봇은 음성체널 안에 없지만, {message.author}이 나가라고 했다')
                    await message.channel.send('`키키야 음성입장`을 입력해 키키봇이 음성체널에 들어가게 만드세요')
            elif message.content.startswith('키키야 음성재생'):
                cut=message.content.split(' ')
                url=cut[2]
                song_there=os.path.isfile('song.mp3')
                try:
                    if song_there:
                        os.remove('song.mp3')
                        print('옛날 음악파일 제거완료')
                except PermissionError:
                    print('음악파일을 제거하려 했지만 재생되고 있네요')
                    await message.channel.send('음악 재생중입니다')
                    return
                await message.channel.send('준비됨')
                voice=get(app.voice_clients, guild=message.guild)
                ydl_ops={
                    'format':'bestaudio/best',
                    'postprocessors':[{
                        'key':'FFmpegExtractAudio',
                        'preferredcodec':'mp3',
                        'preferredquality':'192',
                    }],
                }
                with youtube_dl.YoutubeDL(ydl_ops) as ydl:
                    print('음악 파일 다운로드 중')
                    ydl.download([url])
                for file in os.listdir('./'):
                    if file.endswith('.mp3'):
                        name=file
                        print('음악파일 리네임 완료')
                        os.rename(file, 'song.mp3')
                voice.play(discord.FFmpegPCMAudio('song.mp3'), after=lambda e: print('음악 재생 끝'))
                voice.source=discord.PCMVolumeTransformer(voice.source)
                voice.source.volume=0.20
                nname=name.rsplit('-', 2)
                await message.channel.send(f'{nname} 재생중')
            elif message.content=='키키야 가입':
                file=openpyxl.load_workbook('정보.xlsx')
                sheet=file.active
                i=1
                while sheet['A'+str(i)].value != str(message.author.id):
                    i+=1
                    if sheet['A'+str(i)].value==None:
                        sheet['A'+str(i)].value=str(message.author.id)
                        sheet['B'+str(i)].value=message.author.display_name
                        sheet['C'+str(i)].value=0
                        sheet['D'+str(i)].value=str(0)
                        sheet['E'+str(i)].value=str(0)
                        await message.channel.send('키키봇 서비스 가입 완료')
                        file.save('정보.xlsx')
                        return 
                await message.channel.send('이미 가입되어 있습니다')
                file.save('정보.xlsx')
            elif message.content=='키키야 내돈':
                file = openpyxl.load_workbook("정보.xlsx")
                sheet = file.active
                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if str(sheet["A" + str(i)].value) == str(message.author.id):
                        await message.channel.send(f'{message.author.display_name}님의 돈:  {str(sheet["C" + str(i)].value)}')
                        flag = 1
                        break
                    i += 1
                if flag == 0 : await message.channel.send("`키키야 가입`을 입력해 먼저 가입을 하세요")
            elif message.content=='키키야 올인':
                file=openpyxl.load_workbook('정보.xlsx')
                sheet=file.active
                i=1
                while sheet['A'+str(i)].value!=None:
                    i+=1
                    if str(sheet['A'+str(i)].value)==str(message.author.id):
                        await message.channel.send('성공하면 현재 돈의 2배를 주지만 실패하면 돈은 0원이 됩니다\n확률은 1/2!')
                        rand=random.randrange(1, 3)
                        if rand==1:
                            ohyeah=int(sheet['C'+str(i)].value)
                            sheet['C'+str(i)].value=ohyeah*2
                            await message.channel.send('성공입니다!!!')
                        elif rand==2:
                            sheet['C'+str(i)].value=0
                            await message.channel.send('실패입니다!!!(복사 붙여넣기를 해 어쩔수 없이 느낌표를 달아야하는 키키봇의 슬픈 이야기)')
                        file.save('정보.xlsx')
                        return
                await message.channel.send("`키키야 가입`을 입력해 먼저 가입을 하세요")
            elif message.content.startswith('키키야 언밴'):
                author=message.guild.get_member(int(message.content[7:25]))
                await message.guild.unban(author)
            elif message.content.startswith('키키야 경고'):
                if int(message.author.id)!=647630912795836437:
                    await message.channel.send('관리자안테 해달라고 해\n너는 권한이 없거든')
                    return
                author = message.guild.get_member(int(message.content[7:25]))
                file = openpyxl.load_workbook('경고.xlsx')
                sheet = file.active
                why = str(message.content[28:])
                i = 1
                while True :
                    if sheet["A" + str(i)].value == str(author) :
                        sheet['B' + str(i)].value = int(sheet["B" + str(i)].value) + 1
                        file.save("경고.xlsx")
                        if sheet["B" + str(i)].value == 7:
                            await message.guild.ban(author)
                            await author.send(str(author) + "님은 경고 7회누적으로 서버에서 추방되었습니다.")
                        else:
                            await author.send(str(author) + "님은 경고를 1회 받았습니다")
                            sheet["C" + str(i)].value = why
                        break
                    if sheet["A" + str(i)].value == None:
                        sheet["A" + str(i)].value = str(author)
                        sheet["B" + str(i)].value = 1
                        sheet["C" + str(i)].value = why
                        file.save("경고.xlsx")
                        await author.send(str(author) + "님은 경고를 1회 받았습니다.")
                        break
                    i += 1
            elif message.content=='키키야 돈줘':
                file = openpyxl.load_workbook("정보.xlsx")
                sheet = file.active
                i = 1
                flag = 0
                while sheet["A" + str(i)].value != None:
                    i += 1
                    if str(sheet["A" + str(i)].value) == str(message.author.id):
                        await message.channel.send('1원부터 1000원 까지 아무 양의 돈을 줍니다')
                        rand=random.randrange(1, 1001)
                        ohyeah=int(sheet['C'+str(i)].value)+rand
                        sheet['C'+str(i)].value=""
                        sheet['C'+str(i)].value=str(ohyeah)
                        await message.channel.send(f'**{rand}**원을 줬습니다\n현재 **{message.author.display_name}**님의 돈 액수: {ohyeah}')
                        flag = 1    
                        file.save('정보.xlsx')
                        break
                if flag == 0 : await message.channel.send("`키키야 가입`을 입력해 먼저 가입을 하세요")
            else:
                file = openpyxl.load_workbook("기억.xlsx")
                sheet = file.active
                q=message.content[4:]
                i = 2
                flag = 0
                while sheet["A" + str(i)].value != None:
                    if sheet["A" + str(i)].value == q:
                        await message.channel.send(str(sheet["B" + str(i)].value))
                        flag = 1
                        break
                    i += 1
                if flag == 0 : await message.channel.send("키키봇의 프로그램에는 없는 명령어에요!")
    except Exception as ex:
        await message.channel.send(f'엌 에러닷\n<@647630912795836437>야 봇 관리좀 재대로 해\n에러 내용: {str(ex)}')
async def my_background_task():
    await app.wait_until_ready()
    while not app.is_closed():
        game = discord.Game("키키야 도움말을 입력해 명령어 확인")
        await app.change_presence(status=discord.Status.idle, activity=game)
        await asyncio.sleep(10)
        game = discord.Game(f'{len(app.guilds)}개의 서버에 참여중')
        await app.change_presence(status=discord.Status.idle, activity=game)
        await asyncio.sleep(10)
        game = discord.Game(f'{len(app.users)}명의 유저들과 소통하는중')
        await app.change_presence(status=discord.Status.idle, activity=game)
        await asyncio.sleep(10)
app.loop.create_task(my_background_task())
app.run(token)  
